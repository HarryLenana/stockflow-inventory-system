from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    flash,
    session,
    send_file,
    current_app
)

from database.models import get_connection
from utils.decorators import (
    login_required,
    roles_required
)

from werkzeug.utils import secure_filename

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook

from datetime import datetime

import os
import io
import csv
import shutil
import zipfile

settings_bp = Blueprint("settings", __name__)


# ============================================
# SETTINGS
# ============================================


@settings_bp.route("/settings", methods=["GET", "POST"])
@login_required
@roles_required("Administrator")
def settings():

    if "user" not in session:
        return redirect("/")

    conn = get_connection()
    cursor = conn.cursor()

    # Load current user
    cursor.execute("""
        SELECT *
        FROM users
        WHERE username=?
    """, (session["user"],))

    user = cursor.fetchone()

    # Load system settings
    cursor.execute("""
        SELECT *
        FROM settings
        LIMIT 1
    """)

    settings = cursor.fetchone()

    if request.method == "POST":

        # =====================================
        # COMPANY SETTINGS
        # =====================================

        if request.form.get("company_name") is not None:

            cursor.execute("""
                UPDATE settings
                SET
                    company_name=?,
                    company_email=?,
                    company_phone=?,
                    company_address=?,
                    company_website=?,
                    currency=?,
                    tax_rate=?,
                    low_stock=?
                WHERE id=1
            """, (

                request.form.get("company_name"),
                request.form.get("company_email"),
                request.form.get("company_phone"),
                request.form.get("company_address"),
                request.form.get("company_website"),
                request.form.get("currency"),
                request.form.get("tax_rate"),
                request.form.get("low_stock")

            ))

            conn.commit()

            flash(
                "Settings updated successfully.",
                "success"
            )

            conn.close()

            return redirect("/settings")

        # =====================================
        # PROFILE PICTURE
        # =====================================

        picture = request.files.get("profile_picture")

        if picture and picture.filename != "":

            filename = secure_filename(picture.filename)

            picture.save(
                os.path.join(
                    current_app.config["UPLOAD_FOLDER"],
                    filename
                )
            )

            cursor.execute("""
                UPDATE users
                SET profile_picture=?
                WHERE username=?
            """, (
                filename,
                session["user"]
            ))

            conn.commit()

            flash(
                "Profile picture updated successfully!",
                "success"
            )

            conn.close()

            return redirect("/settings")

        # =====================================
        # CHANGE PASSWORD
        # =====================================

        current_password = request.form.get("current_password")
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")

        if current_password:

            if current_password != user[2]:

                flash(
                    "Current password is incorrect.",
                    "danger"
                )

            elif new_password != confirm_password:

                flash(
                    "Passwords do not match.",
                    "danger"
                )

            else:

                cursor.execute("""
                    UPDATE users
                    SET password=?
                    WHERE username=?
                """, (
                    new_password,
                    session["user"]
                ))

                conn.commit()

                flash(
                    "Password changed successfully.",
                    "success"
                )

            conn.close()

            return redirect("/settings")

    # Reload user

    cursor.execute("""
        SELECT *
        FROM users
        WHERE username=?
    """, (session["user"],))

    user = cursor.fetchone()

    # Reload settings

    cursor.execute("""
        SELECT *
        FROM settings
        LIMIT 1
    """)

    settings = cursor.fetchone()

    conn.close()

    return render_template(
        "settings.html",
        user=user,
        settings=settings
    )


@settings_bp.route("/backup")
def backup_database():

    backup_folder = "backups"
    os.makedirs(backup_folder, exist_ok=True)

    filename = datetime.now().strftime(
        "StockFlow_Backup_%Y-%m-%d_%H-%M-%S.db"
    )

    backup_path = os.path.join(
        backup_folder,
        filename
    )

    shutil.copy(
        "database.db",
        backup_path
    )

    flash("Database backup created successfully!", "success")

    return send_file(
        backup_path,
        as_attachment=True,
        download_name=filename
    )


@settings_bp.route("/restore", methods=["GET", "POST"])
def restore_database():

    if request.method == "POST":

        if "backup_file" not in request.files:
            flash("Please select a backup file.", "danger")
            return redirect("/restore")

        file = request.files["backup_file"]

        if file.filename == "":
            flash("No backup selected.", "danger")
            return redirect("/restore")

        # Save uploaded file temporarily
        temp_file = "temp_restore.db"
        file.save(temp_file)

        # Replace current database
        shutil.copy(temp_file, "database.db")
        os.remove(temp_file)

        flash("Database restored successfully!", "success")
        return redirect("/settings")

    return render_template("restore_database.html")


@settings_bp.route("/export-pdf")
def export_pdf():

    conn = get_connection()
    cursor = conn.cursor()

    # Products
    cursor.execute("""
        SELECT name, category, stock, price
        FROM products
        ORDER BY name
    """)
    products = cursor.fetchall()

    # Revenue
    cursor.execute("""
        SELECT COALESCE(
            SUM(quantity * selling_price),
            0
        )
        FROM sales
    """)
    revenue = cursor.fetchone()[0]

    # Inventory Value
    cursor.execute("""
        SELECT COALESCE(
            SUM(stock * price),
            0
        )
        FROM products
    """)
    inventory = cursor.fetchone()[0]

    conn.close()

    filename = "StockFlow_Report.pdf"

    pdf = canvas.Canvas(filename, pagesize=A4)

    width, height = A4

    y = height - 50

    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawString(50, y, "STOCKFLOW INVENTORY REPORT")

    y -= 30

    pdf.setFont("Helvetica", 12)

    pdf.drawString(50, y, f"Generated: {datetime.now()}")

    y -= 40

    pdf.drawString(50, y, f"Total Revenue: {revenue:,.2f}")

    y -= 20

    pdf.drawString(50, y, f"Inventory Value: {inventory:,.2f}")

    y -= 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, y, "Products")

    y -= 25

    pdf.setFont("Helvetica", 11)

    for product in products:

        line = (
            f"{product[0]} | "
            f"{product[1]} | "
            f"Stock: {product[2]} | "
            f"Price: {product[3]:,.2f}"
        )

        pdf.drawString(50, y, line)

        y -= 18

        if y < 50:
            pdf.showPage()
            y = height - 50

    pdf.save()

    return send_file(
        filename,
        as_attachment=True
    )


@settings_bp.route("/export-excel")
def export_excel():

    conn = get_connection()
    cursor = conn.cursor()

    workbook = Workbook()

    # ==================================
    # PRODUCTS SHEET
    # ==================================
    sheet = workbook.active
    sheet.title = "Products"

    sheet.append(["ID", "Name", "Category", "Stock", "Price"])

    cursor.execute("""
        SELECT id, name, category, stock, price
        FROM products
        ORDER BY name
    """)

    for row in cursor.fetchall():
        sheet.append(list(row))

    # ==================================
    # SUPPLIERS SHEET
    # ==================================
    sheet2 = workbook.create_sheet("Suppliers")

    sheet2.append([
        "ID",
        "Name",
        "Phone",
        "Email",
        "Address"
    ])

    cursor.execute("""
        SELECT id, name, phone, email, address
        FROM suppliers
        ORDER BY name
    """)

    for row in cursor.fetchall():
        sheet2.append(list(row))

    # ==================================
    # PURCHASES SHEET
    # ==================================
    sheet3 = workbook.create_sheet("Purchases")

    sheet3.append([
        "Product",
        "Supplier",
        "Quantity",
        "Cost",
        "Date"
    ])

    cursor.execute("""
        SELECT
            products.name,
            suppliers.name,
            purchases.quantity,
            purchases.cost,
            purchases.purchase_date
        FROM purchases
        JOIN products
            ON purchases.product_id = products.id
        JOIN suppliers
            ON purchases.supplier_id = suppliers.id
    """)

    for row in cursor.fetchall():
        sheet3.append(list(row))

    # ==================================
    # SALES SHEET
    # ==================================
    sheet4 = workbook.create_sheet("Sales")

    sheet4.append([
        "Product",
        "Quantity",
        "Price",
        "Customer",
        "Date"
    ])

    cursor.execute("""
        SELECT
            products.name,
            sales.quantity,
            sales.selling_price,
            sales.customer_name,
            sales.sale_date
        FROM sales
        JOIN products
            ON sales.product_id = products.id
    """)

    for row in cursor.fetchall():
        sheet4.append(list(row))

    # ==================================
    # USERS SHEET
    # ==================================
    sheet5 = workbook.create_sheet("Users")

    sheet5.append([
        "Full Name",
        "Username",
        "Role",
        "Email",
        "Phone"
    ])

    cursor.execute("""
        SELECT
            full_name,
            username,
            role,
            email,
            phone
        FROM users
    """)

    for row in cursor.fetchall():
        sheet5.append(list(row))

    conn.close()

    filename = "StockFlow_Report.xlsx"

    workbook.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )


@settings_bp.route("/export-csv")
def export_csv():

    conn = get_connection()
    cursor = conn.cursor()

    memory_file = io.BytesIO()

    with zipfile.ZipFile(memory_file, "w", zipfile.ZIP_DEFLATED) as zf:

        tables = {
            "products": [
                "id", "name", "category", "stock", "price"
            ],
            "suppliers": [
                "id", "name", "phone", "email", "address"
            ],
            "users": [
                "id", "username", "full_name",
                "email", "phone", "role"
            ]
        }

        for table, headers in tables.items():

            cursor.execute(f"SELECT * FROM {table}")

            rows = cursor.fetchall()

            output = io.StringIO()

            writer = csv.writer(output)

            writer.writerow(headers)

            writer.writerows(rows)

            zf.writestr(
                f"{table}.csv",
                output.getvalue()
            )

        # -----------------------------
        # Purchases
        # -----------------------------

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([
            "Product",
            "Supplier",
            "Quantity",
            "Cost",
            "Date"
        ])

        cursor.execute("""
            SELECT
                products.name,
                suppliers.name,
                purchases.quantity,
                purchases.cost,
                purchases.purchase_date
            FROM purchases
            JOIN products
                ON purchases.product_id = products.id
            JOIN suppliers
                ON purchases.supplier_id = suppliers.id
        """)

        writer.writerows(cursor.fetchall())

        zf.writestr(
            "purchases.csv",
            output.getvalue()
        )

        # -----------------------------
        # Sales
        # -----------------------------

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([
            "Product",
            "Quantity",
            "Price",
            "Customer",
            "Date"
        ])

        cursor.execute("""
            SELECT
                products.name,
                sales.quantity,
                sales.selling_price,
                sales.customer_name,
                sales.sale_date
            FROM sales
            JOIN products
                ON sales.product_id = products.id
        """)

        writer.writerows(cursor.fetchall())

        zf.writestr(
            "sales.csv",
            output.getvalue()
        )

    conn.close()

    memory_file.seek(0)

    return send_file(
        memory_file,
        as_attachment=True,
        download_name="StockFlow_CSV_Export.zip",
        mimetype="application/zip"
    )
